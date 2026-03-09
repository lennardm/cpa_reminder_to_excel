"""
Convert CPA Global patent renewal reminder PDFs to Excel.

Each PDF page contains a table with columns:
  Land | Patent/Ans.nr. | Innehavare | Er referens | År | Förfallodag | Kostnad SEK

Usage:
    python convert.py <input.pdf> [output.xlsx]
    python convert.py Reminder_8248790.pdf
"""

import sys
import re
from pathlib import Path
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Land",
    "Patent/Ans.nr.",
    "Innehavare",
    "Er referens",
    "År",
    "Förfallodag",
    "Kostnad SEK",
]

SWEDISH_MONTHS = {
    "januari", "februari", "mars", "april", "maj", "juni",
    "juli", "augusti", "september", "oktober", "november", "december",
}

# Default column x-boundaries (left edge of each column header, discovered from PDFs).
# Words are assigned to the column whose x-range they fall in.
DEFAULT_COL_BOUNDARIES = [32.8, 126.4, 236.0, 383.5, 531.4, 555.6, 611.5]


def detect_column_boundaries(words):
    """
    Find x-positions of column headers by looking for known header words.
    Returns a list of 7 x-start values, or DEFAULT_COL_BOUNDARIES if not found.
    """
    header_map = {}
    for w in words:
        t = w["text"]
        if t == "Land":
            header_map["Land"] = w["x0"]
        elif t == "Patent/Ans.nr.":
            header_map["Patent"] = w["x0"]
        elif t == "Innehavare":
            header_map["Innehavare"] = w["x0"]
        elif t == "referens":   # "Er referens" split into two words
            header_map["ErRef"] = w["x0"]
        elif t == "År":
            header_map["År"] = w["x0"]
        elif t == "Förfallodag":
            header_map["Förfallodag"] = w["x0"]
        elif "Kostnad" in t or "Kostna" in t:
            header_map["Kostnad"] = w["x0"]

    if len(header_map) >= 6:
        return [
            header_map.get("Land", DEFAULT_COL_BOUNDARIES[0]),
            header_map.get("Patent", DEFAULT_COL_BOUNDARIES[1]),
            header_map.get("Innehavare", DEFAULT_COL_BOUNDARIES[2]),
            header_map.get("ErRef", DEFAULT_COL_BOUNDARIES[3]),
            header_map.get("År", DEFAULT_COL_BOUNDARIES[4]),
            header_map.get("Förfallodag", DEFAULT_COL_BOUNDARIES[5]),
            header_map.get("Kostnad", DEFAULT_COL_BOUNDARIES[6]),
        ]
    return DEFAULT_COL_BOUNDARIES


def assign_column(x0, boundaries):
    """
    Return column index (0-6) based on x position.
    Uses header position minus a left-tolerance so data that starts slightly
    left of the header label still lands in the correct column.
    """
    LEFT_TOLERANCE = 15  # px — data can start this far left of the header
    col = 0
    for i, boundary in enumerate(boundaries):
        if x0 >= boundary - LEFT_TOLERANCE:
            col = i
    return col


def group_words_by_line(words, y_tolerance=3):
    """Group words into lines by their top y-coordinate."""
    lines = {}
    for w in words:
        y = round(w["top"] / y_tolerance) * y_tolerance
        lines.setdefault(y, []).append(w)
    return {y: sorted(ws, key=lambda w: w["x0"]) for y, ws in lines.items()}


def looks_like_reference(s):
    return bool(s and re.match(r"^P\d", s))


def looks_like_date(s):
    return bool(s and any(m in s.lower() for m in SWEDISH_MONTHS))


def extract_table_rows(pdf_path):
    all_rows = []
    meta = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)

            if page_num == 0:
                text = page.extract_text() or ""
                m = re.search(r"Kundnr[.\s]*\n?\s*(\d+)", text)
                if m:
                    meta["kundnr"] = m.group(1)
                m = re.search(r"\b(\d{2}\s+\w+\s+\d{4})\b", text)
                if m:
                    meta["datum"] = m.group(1)
                m = re.search(r"\b(8\d{6})\b", text)
                if m:
                    meta["doc_nr"] = m.group(1)

            boundaries = detect_column_boundaries(words)
            lines = group_words_by_line(words)

            for y, line_words in sorted(lines.items()):
                # Build columns from word positions
                cols = {i: [] for i in range(7)}
                for w in line_words:
                    col_idx = assign_column(w["x0"], boundaries)
                    cols[col_idx].append(w["text"])

                land = " ".join(cols[0])
                patent = " ".join(cols[1])
                innehavare = " ".join(cols[2])
                er_ref = " ".join(cols[3])
                ar = " ".join(cols[4])
                forfallo = " ".join(cols[5])
                kostnad = " ".join(cols[6])

                # A data row must have: non-empty land, patent, valid er_ref, valid date
                if (
                    land and patent
                    and looks_like_reference(er_ref)
                    and looks_like_date(forfallo)
                ):
                    all_rows.append([land, patent, innehavare, er_ref, ar, forfallo, kostnad])

    return meta, all_rows


def write_excel(rows, meta, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Patent Renewals"

    # --- Metadata header block ---
    ws["A1"] = "CPA Global – Årsavgifter för patent/patentansökningar"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")

    row_idx = 2
    if meta.get("kundnr"):
        ws.cell(row_idx, 1, f"Kundnr: {meta['kundnr']}")
    if meta.get("datum"):
        ws.cell(row_idx, 3, f"Datum: {meta['datum']}")
    if meta.get("doc_nr"):
        ws.cell(row_idx, 5, f"Dokument: {meta['doc_nr']}")
    row_idx += 2

    # --- Column headers ---
    header_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row_idx, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    row_idx += 1

    data_start_row = row_idx

    # --- Data rows ---
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    for i, row in enumerate(rows):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.border = border
            cell.fill = fill
            if col == 5:  # År
                try:
                    cell.value = float(value) if "." in value else int(value)
                    cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    pass
            elif col == 7:  # Kostnad SEK
                try:
                    cell.value = int(value)
                    cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    pass
            elif col == 6:  # Förfallodag
                cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    # --- Total row ---
    ws.cell(row_idx, 6, "Totalt").font = Font(bold=True)
    ws.cell(row_idx, 6).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row_idx, 7)
    total_cell.value = f"=SUM(G{data_start_row}:G{row_idx - 1})"
    total_cell.font = Font(bold=True)
    total_cell.border = border

    # --- Column widths ---
    col_widths = [18, 22, 30, 16, 6, 16, 14]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze panes below header
    ws.freeze_panes = f"A{data_start_row}"

    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(rows)} rows)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python convert.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    if not pdf_path.exists():
        print(f"File not found: {pdf_path}")
        sys.exit(1)

    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else pdf_path.with_suffix(".xlsx")

    print(f"Extracting from {pdf_path} ...")
    meta, rows = extract_table_rows(pdf_path)
    print(f"Found {len(rows)} data rows. Metadata: {meta}")

    if not rows:
        print("No data rows found. Check PDF structure.")
        sys.exit(1)

    write_excel(rows, meta, output_path)


if __name__ == "__main__":
    main()
